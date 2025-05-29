import os
import re
import io
import pandas as pd
import streamlit as st
from PIL import Image
# --------------------------------------
# Page config and CSS
# --------------------------------------
st.set_page_config(page_title="College Predictor", layout="wide")
st.markdown("""
<style>
.card-item { 
    background: #fff; 
    border-radius: 8px; 
    box-shadow: 0 2px 8px rgba(0,0,0,0.1); 
    width: 358px; 
    overflow: hidden; 
    margin-bottom: 16px; 
}
.card-img  { 
    width: 100%; 
    height: 130px; 
    object-fit: cover; 
}
.card-body { 
    padding: 12px; 
}
.card-title{ 
    font-size: 16px; 
    font-weight: 600; 
    margin-bottom: 4px; 
}
.card-univ { 
    font-size: 14px; 
    color: #555; 
    margin-bottom: 8px; 
}
.card-footer{ 
    font-size: 13px; 
    color: #333; 
    margin-bottom: 8px; 
}
</style>
""", unsafe_allow_html=True)

# --------------------------------------
# Constants & Helpers
# --------------------------------------
GRID_SIZE = (358, 130)
PAGE_SIZE = 10

def format_inr(val):
    if pd.isna(val):
        return "-"
    s = str(int(val))
    r, s = s[-3:], s[:-3]
    while s:
        r = s[-2:] + "," + r
        s = s[:-2]
    return f"₹{r}"

def slugify(name: str) -> str:
    return re.sub(r'[^a-z0-9]+', '_', name.lower()).strip('_')

def load_and_resize(path, size):
    try:
        img = Image.open(path)
        return img.resize(size, Image.Resampling.LANCZOS)
    except Exception:
        return None

# --------------------------------------
# Load Data
# --------------------------------------
@st.cache_data
def load_data(path):
    df = pd.read_csv(path)
    # Clean up column names
    df.columns = (
        df.columns.astype(str)
          .str.replace(r'"','', regex=True)
          .str.replace(r'[\r\n]+',' ', regex=True)
          .str.strip()
          .str.replace(r'\s+','_', regex=True)
          .str.upper()
    )
    df = df.rename(columns={
        'UNIVESITY_NAME': 'UNIVERSITY_NAME',
        'HOSTEL_CHARGES_P.A.': 'HOSTEL_CHARGES_PA'
    })
    # Clean up numeric columns
    for col in ['TUITION_FEE', 'GRAND_TOTAL', 'OPENING_RANK_2023', 'CLOSING_RANK_2023']:
        if col in df:
            df[col] = df[col].astype(str).str.replace(r'[₹,]', '', regex=True).str.strip()
            df[col] = pd.to_numeric(df[col], errors='coerce')
    return df

df = load_data("college_data_predictor.csv")

# --------------------------------------
# Session defaults
# --------------------------------------
ss = st.session_state
for k, v in {
    'search_done': False,
    'selected_idx': None,
    'page': 1,
    'prev_sel': []
}.items():
    if k not in ss:
        ss[k] = v
if 'selected_colleges' not in ss:
    ss.selected_colleges = []

# --------------------------------------
# Excel styling & download helper
# --------------------------------------
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment

def style_and_download(df_export, label, filename, sheet_name="Sheet1"):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name=sheet_name)
        wb  = writer.book
        ws  = wb[sheet_name]
        cols = df_export.columns.tolist()

        # 1) Header styling
        light_green = PatternFill("solid", fgColor="C6EFCE")
        dark_green  = PatternFill("solid", fgColor="006100")
        for idx, cell in enumerate(ws[1], start=1):
            cell.font = Font(bold=True, color="000000")
            if idx <= cols.index("Website")+1:
                cell.fill = light_green
            else:
                cell.fill = dark_green
                cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        # 2) Alternate row coloring
        alt = PatternFill("solid", fgColor="F2F2F2")
        for r in range(2, ws.max_row+1):
            if r % 2 == 0:
                for c in range(1, ws.max_column+1):
                    ws.cell(r,c).fill = alt

        # 3) Budget Status coloring
        bs_col = cols.index("Budget Status") + 1
        red    = PatternFill("solid", fgColor="FFC7CE")
        green  = PatternFill("solid", fgColor="C6EFCE")
        for r in range(2, ws.max_row+1):
            cell = ws.cell(r, bs_col)
            cell.fill = red if cell.value == "Budget Exceeding" else green

        # 4) Rupee formatting
        rupee_fmt = u'₹#,##0'
        for fee_col in ["Tuition fee", "Grand Total"]:
            if fee_col in cols:
                ci = cols.index(fee_col) + 1
                for r in range(2, ws.max_row+1):
                    ws.cell(r, ci).number_format = rupee_fmt

        # 5) Auto-width
        for col_cells in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max_len + 2

        # no writer.save() needed with pandas >= 1.4
    buf.seek(0)
    st.download_button(label, buf, filename,
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# --------------------------------------
# Search Form
# --------------------------------------
if not ss.search_done:
    st.title("🎓 College Predictor")
    st.markdown("Enter your NEET AIR and budgets to see qualified colleges.")
    with st.form("search_form"):
        ur = st.number_input("NEET AIR (Rank):", min_value=1, value=1)
        tb = st.number_input("Tuition budget (₹):", min_value=0, value=2_000_000)
        ob = st.number_input("Overall budget (₹):", min_value=0, value=3_000_000)
        if st.form_submit_button("Search"):
            ss.search_done    = True
            ss.user_rank      = ur
            ss.tuition_budget = tb
            ss.overall_budget = ob
            ss.page           = 1
            ss.prev_sel       = []
            
            # Modified eligibility logic:
            def compute(r):
                cl = r['CLOSING_RANK_2023']
                tf, gt = r['TUITION_FEE'], r['GRAND_TOTAL']
                # If the closing rank is missing or the user's rank is worse than the closing rank, it's not possible.
                if pd.isna(cl) or (ur > cl):
                    return "NOT POSSIBLE"
                # Check if tuition fee or overall cost exceeds the budget.
                if (pd.notna(tf) and tf > tb) or (pd.notna(gt) and gt > ob):
                    return "Budget Exceeding"
                return "Within budget"
            
            df['Budget Status'] = df.apply(compute, axis=1)
            ss.possible_colleges = df[df['Budget Status'].isin(["Within budget", "Budget Exceeding"])].reset_index(drop=True)
    st.stop()

possible_colleges = ss.possible_colleges

# --------------------------------------
# Modify Search
# --------------------------------------
if st.button("Modify Search"):
    ss.search_done    = False
    ss.selected_idx   = None
    ss.selected_colleges.clear()
    for k in list(st.session_state):
        if k.startswith("select_"):
            del st.session_state[k]
    ss.page = 1
    st.rerun()

# --------------------------------------
# State Filter
# --------------------------------------
selected_states = []
if 'STATE' in possible_colleges.columns:
    cnts = possible_colleges['STATE'].value_counts().to_dict()
    opts = [f"{s} ({cnts[s]})" for s in sorted(cnts)]
    sel = st.multiselect("Filter by state:", opts, default=ss.prev_sel)
    if sel:
        selected_states = [s.split(" (")[0] for s in sel]
        results = possible_colleges[possible_colleges['STATE'].isin(selected_states)].reset_index(drop=True)
    else:
        results = possible_colleges.reset_index(drop=True)
    if sel != ss.prev_sel:
        ss.page     = 1
        ss.prev_sel = sel
else:
    results = possible_colleges.reset_index(drop=True)

# --------------------------------------
# Pagination
# --------------------------------------
num_pages = len(results) // PAGE_SIZE + int(len(results) % PAGE_SIZE != 0)
def paginate(direction):
    if direction == "prev" and ss.page > 1:
        ss.page -= 1
    if direction == "next" and ss.page < num_pages:
        ss.page += 1

start_idx = (ss.page - 1) * PAGE_SIZE
page_data = results.iloc[start_idx:start_idx + PAGE_SIZE].reset_index(drop=True)

# --------------------------------------
# Summary
# --------------------------------------
ur, tb, ob = ss.user_rank, ss.tuition_budget, ss.overall_budget
base = f"As per your rank {ur} and budgets {format_inr(tb)}/{format_inr(ob)}, "
if selected_states:
    st.markdown(base + f"you qualify for **{len(results)}** colleges in **{', '.join(selected_states)}**.")
else:
    st.markdown(base + f"you qualify for **{len(possible_colleges)}** colleges.")
st.markdown(base + "…")

# --------------------------------------
# Comparison mode
# --------------------------------------
compare_choices = st.multiselect(
    "🔍 Compare two colleges:",
    options=possible_colleges["COLLEGE"],
    default=st.session_state.get("compare_sel", []),
    help="Select exactly two colleges to see their stats side‑by‑side"
)
st.session_state["compare_sel"] = compare_choices

if len(compare_choices) == 2:
    # Filter the data for the selected colleges and remove duplicate entries
    comp_df = possible_colleges[possible_colleges["COLLEGE"].isin(compare_choices)]
    comp_df = comp_df.drop_duplicates(subset=["COLLEGE"])
    comp_df = comp_df[[
        "COLLEGE", "UNIVERSITY_NAME", "STATE",
        "TUITION_FEE", "HOSTEL_CHARGES_PA", "ANNUAL",
        "TUITION_PACKAGE", "GRAND_TOTAL",
        "OPENING_RANK_2023", "CLOSING_RANK_2023", "Budget Status"
    ]].rename(columns={
        "COLLEGE": "College", "UNIVERSITY_NAME": "University",
        "TUITION_FEE": "Tuition Fee", "HOSTEL_CHARGES_PA": "Hostel P.A.",
        "GRAND_TOTAL": "Grand Total", "OPENING_RANK_2023": "Open Rank 2023",
        "CLOSING_RANK_2023": "Close Rank 2023"
    }).set_index("College").T
    st.markdown("#### Comparison")
    st.dataframe(comp_df)
elif len(compare_choices) > 2:
    st.warning("Please select only **two** colleges for comparison.")

# --------------------------------------
# Persist selections via unified list
# --------------------------------------
st.markdown(f"**Selected: {len(ss.selected_colleges)}**")

# --------------------------------------
# Client name prompt
# --------------------------------------
client_name = st.text_input("Enter client name (for sheet)")

# --------------------------------------
# Download Selected (PDF)
# --------------------------------------
# --------------------------------------
# Download Selected (Excel only)
# --------------------------------------
if ss.selected_colleges:
    df_sel = possible_colleges[possible_colleges['COLLEGE'].isin(ss.selected_colleges)].copy()
    export_sel = df_sel[[
        "COLLEGE", "UNIVERSITY_NAME", "STATE", "SEATS", "OVERVIEW", "WEBSITE",
        "TUITION_FEE", "HOSTEL_CHARGES_PA", "ANNUAL", "ONE_TIME",
        "TUITION_PACKAGE", "GRAND_TOTAL", "Budget Status"
    ]].rename(columns={
        "COLLEGE": "College", "UNIVERSITY_NAME": "University", "STATE": "State",
        "SEATS": "Seats", "OVERVIEW": "Overview", "WEBSITE": "Website",
        "TUITION_FEE": "Tuition fee", "HOSTEL_CHARGES_PA": "Hostel charges",
        "ANNUAL": "Annual", "ONE_TIME": "One Time",
        "TUITION_PACKAGE": "Tuition Package", "GRAND_TOTAL": "Grand Total"
    })
    if not client_name.strip():
        st.warning("Please enter a client name before downloading.")
    else:
        style_and_download(
            export_sel,
            label=f"📥 Download Selected for {client_name}",
            filename=f"{client_name}_selected.xlsx",
            sheet_name=client_name
        )
else:
    st.info("Select one or more colleges to enable download.")

# --------------------------------------
# Clear & Download All (Excel only)
# --------------------------------------
c1, c2 = st.columns(2)
with c1:
    if st.button("🗑️ Clear Selections"):
        ss.selected_colleges.clear()
        for k in list(st.session_state):
            if k.startswith("select_"):
                del st.session_state[k]
        ss.page = 1
        st.experimental_rerun()
with c2:
    export_all = results[[
        "COLLEGE", "UNIVERSITY_NAME", "STATE", "SEATS", "OVERVIEW", "WEBSITE",
        "TUITION_FEE", "HOSTEL_CHARGES_PA", "ANNUAL", "ONE_TIME",
        "TUITION_PACKAGE", "GRAND_TOTAL", "Budget Status"
    ]].rename(columns={
        "COLLEGE": "College", "UNIVERSITY_NAME": "University", "STATE": "State",
        "SEATS": "Seats", "OVERVIEW": "Overview", "WEBSITE": "Website",
        "TUITION_FEE": "Tuition fee", "HOSTEL_CHARGES_PA": "Hostel charges",
        "ANNUAL": "Annual", "ONE_TIME": "One Time",
        "TUITION_PACKAGE": "Tuition Package", "GRAND_TOTAL": "Grand Total"
    })
    if not client_name.strip():
        st.warning("Please enter a client name before downloading.")
    else:
        style_and_download(
            export_all,
            label=f"📥 Download All for {client_name}",
            filename=f"{client_name}_all.xlsx",
            sheet_name=client_name
        )

# --------------------------------------
# Detail view
# --------------------------------------
if ss.selected_idx is not None and ss.selected_idx < len(results):
    row = results.loc[ss.selected_idx]
    st.markdown(f"### {row['COLLEGE'].split(',')[0]}")
    st.markdown(f"**{row['UNIVERSITY_NAME']}**")
    img = row.get('IMAGE_PATH', f"images/{slugify(row['COLLEGE'])}.jpg")
    pil = load_and_resize(img, GRID_SIZE)
    st.image(pil or "https://via.placeholder.com/358x130", width=GRID_SIZE[0])
    st.markdown(f"**Tuition Fee:** {format_inr(row['TUITION_FEE'])} • **Grand Total:** {format_inr(row['GRAND_TOTAL'])}")
    st.markdown(f"**Opening Rank 2023:** {int(row['OPENING_RANK_2023'])} • **Closing Rank 2023:** {int(row['CLOSING_RANK_2023'])}")
    st.markdown(f"**Budget Status:** {row['Budget Status']}")
    st.write(row.get('OVERVIEW', 'No overview available.'))
    if row.get('WEBSITE'):
        st.markdown(f"[Visit Website]({row['WEBSITE']})")
    if st.button("← Back to Results"):
        ss.selected_idx = None
        st.experimental_rerun()

# --------------------------------------
# Grid of cards
# --------------------------------------
st.markdown("## Available Colleges")
cols = st.columns(3, gap="large")
for i, row in page_data.iterrows():
    idx = start_idx + i
    with cols[i % 3]:
        st.markdown("<div class='card-item'>", unsafe_allow_html=True)
        img = row.get('IMAGE_PATH', f"images/{slugify(row['COLLEGE'])}.jpg")
        pil = load_and_resize(img, GRID_SIZE)
        st.image(pil or "https://via.placeholder.com/358x130", width=GRID_SIZE[0])
        st.markdown("<div class='card-body'>", unsafe_allow_html=True)
        st.markdown(f"**{row['COLLEGE'].split(',')[0]}**")
        st.markdown(f"*{row['UNIVERSITY_NAME']}*")
        line = f"Tuition: {format_inr(row['TUITION_FEE'])}"
        if pd.notna(row.get('STATE')):
            line += f" | State: {row['STATE']}"
        if pd.notna(row.get('CLOSING_RANK_2023')):
            line += f" | Closing: {int(row['CLOSING_RANK_2023'])}"
        line += f" | Budget: {row['Budget Status']}"
        st.markdown(f"<div class='card-footer'>{line}</div>", unsafe_allow_html=True)
        st.markdown("</div></div>", unsafe_allow_html=True)

        # Persistent checkbox logic for selection
        slug = slugify(row['COLLEGE'])
        checked = st.checkbox("Select", value=(row['COLLEGE'] in ss.selected_colleges), key=f"select_{slug}")
        if checked and row['COLLEGE'] not in ss.selected_colleges:
            ss.selected_colleges.append(row['COLLEGE'])
        if not checked and row['COLLEGE'] in ss.selected_colleges:
            ss.selected_colleges.remove(row['COLLEGE'])

        if st.button("View Details", key=f"view_{idx}"):
            ss.selected_idx = idx
            st.rerun()

# --------------------------------------
# Pagination controls
# --------------------------------------
# --------------------------------------
# Pagination controls
# --------------------------------------
c1, c2, c3 = st.columns([1, 2, 1])
with c1:
    if st.button("⬅ Previous Page") and ss.page > 1:
        paginate("prev")
with c3:
    if st.button("Next Page ➡") and ss.page < num_pages:
        paginate("next")
with c2:
    st.markdown(
        f"<p style='text-align:center;'>Page {ss.page} of {num_pages}</p>",
        unsafe_allow_html=True
    )
